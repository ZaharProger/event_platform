from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework import status
from rest_framework.permissions import IsAuthenticated

import openpyxl
from openpyxl.styles import Font, Alignment, Color, PatternFill, Border
import os
from datetime import datetime

from django.http import HttpResponse
from django.db.transaction import atomic

from tasks.models import Task, UserTask
from .models import DocField, Doc, FieldValue
from .forms import FieldValueForm
from .serializers import FieldValueSerializer
from events.models import Event
from users.models import UserPassport


class DocsView(APIView):
    authentication_classes = [SessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def put(self, request):
        found_passport = UserPassport.objects.filter(username=request.user.username)
        found_event = Event.objects.filter(pk=request.data['event_id'])
        found_doc = Doc.objects.filter(pk=request.data['doc_id'])

        if len(found_passport) != 0 and len(found_event) != 0 and len(found_doc) != 0:
            with atomic():
                found_doc[0].name = request.data['name']
                found_doc[0].save()

                request_pairs = []
                for field in request.data['fields']:
                    request_pairs.extend([(field['id'], value) for \
                                            value in field['values']])
                request_values = [pair[1]['id'] for pair in request_pairs \
                                  if type(pair[1]['id']) != str]
                for value in FieldValue.objects.filter(pk__in=request_values):
                    if value.pk not in request_values:
                        value.delete()
                
                for pair in request_pairs:
                    found_value = [] if type(pair[1]['id']) == str \
                        else FieldValue.objects.filter(pk=pair[1]['id'])
                    
                    if len(found_value) != 0:
                        value_data = FieldValueSerializer(found_value[0], data=pair[1])
                    else:
                        value_data = FieldValueForm(pair[1])
                    
                    if value_data.is_valid():
                        added_value = value_data.save()
                        if len(found_value) == 0:
                            found_field = [field for field in found_doc[0].fields.all() \
                                           if field.pk == pair[0]]
                            if len(found_field) != 0:
                                added_value.field = found_field[0]
                                added_value.save()

        return Response(
            {'message': ''},
            status=status.HTTP_200_OK,
            content_type='application/json'
        )

    def get(self, request):
        found_passport = UserPassport.objects.filter(username=request.user.username)
        found_doc = Doc.objects.filter(pk=request.GET.get('id', -1))
        data = None

        if len(found_passport) != 0 and len(found_doc) != 0:
            docs_path = os.path.join(
                'event_platform', 
                'static', 
                found_passport[0].doc_template
            )
            file = [url for url in os.listdir(docs_path) 
                    if found_doc[0].doc_type == url.split('.')[0]]
            
            if len(file) != 0:
                workbook = openpyxl.load_workbook(os.path.join(docs_path, file[0]))
                doc_sheet = workbook.active

                doc_sheet.cell(row=1, column=2).value = found_doc[0].name
                doc_sheet.cell(row=1, column=2).font = Font(size=18.0, color='ffffff')
                center_alignment = Alignment(horizontal='center', vertical='center')

                state_colors = {
                    Task.TaskStates.NOT_ASSIGNED: 'eeeeee',
                    Task.TaskStates.ACTIVE: 'ffa726',
                    Task.TaskStates.COMPLETED: '4caf50'
                }
                format_template = '%d.%m.%Y %H:%M'

                prepared_tasks = []
                parent_tasks = Task.objects \
                    .filter(event=found_doc[0].event, parent=None) \
                    .order_by('datetime_start')
                
                for parent_task in parent_tasks:
                    nested_tasks = Task.objects \
                        .filter(parent=parent_task.pk) \
                        .order_by('datetime_start')
                    
                    tasks_group = [parent_task]
                    tasks_group.extend(nested_tasks)
                    prepared_tasks.append(tasks_group)

                doc_row = 3
                for i in range(len(prepared_tasks)):
                    for j in range(len(prepared_tasks[i])):
                        doc_row += 1

                        datetime_start = prepared_tasks[i][j].datetime_start
                        datetime_end = prepared_tasks[i][j].datetime_end
                        formatted_datetime_start = datetime \
                            .fromtimestamp(datetime_start) \
                            .strftime(format_template)
                        task_dates = f'{formatted_datetime_start}'

                        if prepared_tasks[i][j].datetime_end is not None:
                            formatted_datetime_end = datetime \
                                .fromtimestamp(datetime_end) \
                                .strftime(format_template)
                            task_dates += f' - {formatted_datetime_end}'
                    
                        responsible_user = UserTask.objects.filter(
                            task=prepared_tasks[i][j], 
                            is_responsible=True
                        )
                        responsible_user = '' if len(responsible_user) == 0 \
                            else responsible_user[0].user.name
                    
                        task_users = ', '.join([task_user.name for task_user \
                                                in prepared_tasks[i][j].users.all()])

                        if j == 0:
                            item_number = f'{i + 1}'
                            cell_font = Font(bold=True, size=14.0)
                            cell_color = PatternFill(
                                patternType='solid',
                                start_color=Color('aaaaaa')
                            )
                        else:
                            item_number = f'{i + 1}.{j}'
                            cell_font = Font(bold=True, size=12.0)
                            cell_color = PatternFill(
                                patternType='solid',
                                start_color=Color('cccccc')
                            )

                        doc_sheet.cell(row=doc_row, column=1).value = item_number
                        doc_sheet.cell(row=doc_row, column=1).font = cell_font
                        doc_sheet.cell(row=doc_row, column=1).fill = cell_color

                        doc_sheet.cell(row=doc_row, column=2).value = prepared_tasks[i][j].name
                        doc_sheet.cell(row=doc_row, column=2).font = cell_font
                        doc_sheet.cell(row=doc_row, column=2).fill = cell_color

                        doc_sheet.cell(row=doc_row, column=3).value = task_dates
                        doc_sheet.cell(row=doc_row, column=3).alignment = center_alignment
                        doc_sheet.cell(row=doc_row, column=3).font = cell_font
                        doc_sheet.cell(row=doc_row, column=3).fill = cell_color

                        doc_sheet.cell(row=doc_row, column=4).value = prepared_tasks[i][j].state
                        doc_sheet.cell(row=doc_row, column=4).font = cell_font
                        doc_sheet.cell(row=doc_row, column=4).fill = PatternFill(
                            patternType='solid',
                            start_color=Color(state_colors[prepared_tasks[i][j].state])
                        )
                        doc_sheet.cell(row=doc_row, column=5).value = responsible_user
                        doc_sheet.cell(row=doc_row, column=5).font = cell_font
                        doc_sheet.cell(row=doc_row, column=5).fill = cell_color

                        doc_sheet.cell(row=doc_row, column=6).value = task_users
                        doc_sheet.cell(row=doc_row, column=6).font = cell_font
                        doc_sheet.cell(row=doc_row, column=6).fill = cell_color

                export_path = os.path.join(
                    'event_platform', 
                    'static', 
                    'export', 
                    f'{found_doc[0].name}.xlsx'
                )
                workbook.save(export_path)
                workbook.close()
                
            with open(os.path.join(export_path), 'rb') as created_file:
                data = created_file.read()
            os.remove(export_path)

        response = HttpResponse(
            data,
            status=status.HTTP_200_OK if data is not None else status.HTTP_404_NOT_FOUND,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        if data is not None:
            response['Content-Disposition'] = f'attachment; filename="{found_doc[0].name}.xlsx"'

        return response
