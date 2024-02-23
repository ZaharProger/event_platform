from abc import ABC, abstractmethod

import openpyxl
from openpyxl.styles import Font, Alignment, Color, PatternFill, Border
import os
from datetime import datetime, timezone, timedelta

from docs.models import Doc, DocField, FieldValue
from tasks.models import Task, UserTask


class DocBuilder(ABC):
    @abstractmethod
    def build(self, doc: Doc, docs_path: str, export_path: str, file_name: str):
        ...


class TableDocBuilder(DocBuilder):
    def __init__(self, builder_schema=None):
        self.shema = builder_schema

    def build(self, doc: Doc, docs_path: str, export_path: str, file_name: str):
        workbook = openpyxl.load_workbook(os.path.join(docs_path, file_name))
        doc_sheet = workbook.active
        
        doc_sheet.cell(row=1, column=2).value = doc.name
        doc_sheet.cell(row=1, column=2).font = Font(size=18.0, color='ffffff')

        if self.shema is not None:
            self.shema.apply(doc_sheet, doc)
        
        workbook.save(export_path)
        workbook.close()


class BuilderShema:
    def apply(self, doc_sheet, doc: Doc):
        center_alignment = Alignment(horizontal='center', vertical='center')
        end_alignment = Alignment(horizontal='right', vertical='center')
        format_template = '%d.%m.%Y %H:%M'
        header_font = Font(bold=True, size=16.0)
        cell_font = Font(bold=True, size=14.0)
        cell_color = PatternFill(
            patternType='solid',
            start_color=Color('aaaaaa')
        )

        doc_fields = DocField.objects.filter(doc=doc)
        field_values = {field: FieldValue.objects.filter(field=field) for field in doc_fields}

        doc_col = 1
        for field in field_values.keys():
            doc_col += 1
            doc_sheet.cell(row=3, column=doc_col).value = field.name
            doc_sheet.cell(row=3, column=doc_col).font = header_font
        
        for i in range(len(FieldValue.objects.filter(field=list(field_values.keys())[0]))):
            doc_row = 4 + i
            doc_sheet.cell(row=doc_row, column=1).value = f'{i + 1}'
            doc_sheet.cell(row=doc_row, column=1).alignment = center_alignment
            doc_sheet.cell(row=doc_row, column=1).font = cell_font
            doc_sheet.cell(row=doc_row, column=1).fill = cell_color 

        doc_col = 1
        for field, values in field_values.items():
            doc_col += 1
            doc_row = 3

            for value in values:
                doc_row += 1

                if field.field_type == DocField.FieldTypes.NUMBER:
                    doc_sheet.cell(row=doc_row, column=doc_col).alignment = end_alignment
                else:
                    doc_sheet.cell(row=doc_row, column=doc_col).alignment = center_alignment
                
                if field.field_type == DocField.FieldTypes.DATE:
                    field_value = datetime \
                        .fromtimestamp(int(value.value), tz=timezone(timedelta(hours=8))) \
                        .strftime(format_template)
                else:
                    field_value = value.value
                
                doc_sheet.cell(row=doc_row, column=doc_col).value = field_value
                doc_sheet.cell(row=doc_row, column=doc_col).font = cell_font
                doc_sheet.cell(row=doc_row, column=doc_col).fill = cell_color   


class RoadmapSchema(BuilderShema):
    def apply(self, doc_sheet, doc: Doc):
        center_alignment = Alignment(horizontal='center', vertical='center')
        format_template = '%d.%m.%Y %H:%M'

        state_colors = {
            Task.TaskStates.NOT_ASSIGNED: 'eeeeee',
            Task.TaskStates.ACTIVE: 'ffa726',
            Task.TaskStates.COMPLETED: '4caf50'
        }

        prepared_tasks = []
        parent_tasks = Task.objects \
            .filter(event=doc.event, parent=None) \
            .order_by('datetime_start')
                
        for parent_task in parent_tasks:
            nested_tasks = Task.objects \
                .filter(parent=parent_task.pk) \
                .order_by('datetime_start')
                    
            tasks_group = [parent_task]
            tasks_group.extend(nested_tasks)
            prepared_tasks.append(tasks_group)

            doc_row = 3
            for i in range(len(prepared_tasks)):
                for j in range(len(prepared_tasks[i])):
                    doc_row += 1

                    datetime_start = prepared_tasks[i][j].datetime_start
                    datetime_end = prepared_tasks[i][j].datetime_end
                
                    formatted_datetime_start = datetime \
                        .fromtimestamp(datetime_start, tz=timezone(timedelta(hours=8))) \
                        .strftime(format_template)
                    task_dates = f'{formatted_datetime_start}'

                    if prepared_tasks[i][j].datetime_end is not None:
                        formatted_datetime_end = datetime \
                            .fromtimestamp(datetime_end, tz=timezone(timedelta(hours=8))) \
                            .strftime(format_template)
                        task_dates += f' - {formatted_datetime_end}'
                    
                    responsible_user = UserTask.objects.filter(
                        task=prepared_tasks[i][j], 
                        is_responsible=True
                    )
                    responsible_user = '' if len(responsible_user) == 0 \
                        else responsible_user[0].user.name
                    
                    task_users = ', '.join([task_user.name for task_user \
                                            in prepared_tasks[i][j].users.all()])

                    if j == 0:
                        item_number = f'{i + 1}'
                        cell_font = Font(bold=True, size=14.0)
                        cell_color = PatternFill(
                            patternType='solid',
                            start_color=Color('aaaaaa')
                        )
                    else:
                        item_number = f'{i + 1}.{j}'
                        cell_font = Font(bold=True, size=12.0)
                        cell_color = PatternFill(
                            patternType='solid',
                            start_color=Color('cccccc')
                        )

                    doc_sheet.cell(row=doc_row, column=1).value = item_number
                    doc_sheet.cell(row=doc_row, column=1).alignment = center_alignment
                    doc_sheet.cell(row=doc_row, column=1).font = cell_font
                    doc_sheet.cell(row=doc_row, column=1).fill = cell_color

                    doc_sheet.cell(row=doc_row, column=2).value = prepared_tasks[i][j].name
                    doc_sheet.cell(row=doc_row, column=2).font = cell_font
                    doc_sheet.cell(row=doc_row, column=2).fill = cell_color

                    doc_sheet.cell(row=doc_row, column=3).value = task_dates
                    doc_sheet.cell(row=doc_row, column=3).alignment = center_alignment
                    doc_sheet.cell(row=doc_row, column=3).font = cell_font
                    doc_sheet.cell(row=doc_row, column=3).fill = cell_color

                    doc_sheet.cell(row=doc_row, column=4).value = prepared_tasks[i][j].state
                    doc_sheet.cell(row=doc_row, column=4).font = cell_font
                    doc_sheet.cell(row=doc_row, column=4).fill = PatternFill(
                        patternType='solid',
                        start_color=Color(state_colors[prepared_tasks[i][j].state])
                    )
                    doc_sheet.cell(row=doc_row, column=5).value = responsible_user
                    doc_sheet.cell(row=doc_row, column=5).font = cell_font
                    doc_sheet.cell(row=doc_row, column=5).fill = cell_color

                    doc_sheet.cell(row=doc_row, column=6).value = task_users
                    doc_sheet.cell(row=doc_row, column=6).font = cell_font
                    doc_sheet.cell(row=doc_row, column=6).fill = cell_color


class MoneySchema(BuilderShema):
    def apply(self, doc_sheet, doc):
        super().apply(doc_sheet, doc)

        end_alignment = Alignment(horizontal='right', vertical='center')
        cell_font = Font(bold=True, size=14.0)
        cell_color = PatternFill(
            patternType='solid',
            start_color=Color('aaaaaa')
        )
        
        total_sum = 0.0
        doc_field = DocField.objects.filter(doc=doc)[0]
        field_values = FieldValue.objects.filter(field=doc_field)

        for i in range(len(field_values)):
            doc_row = 4 + i
            amount = float(doc_sheet.cell(row=doc_row, column=4).internal_value)
            price = float(doc_sheet.cell(row=doc_row, column=5).internal_value)
    
            cell_sum = amount * price
            total_sum += cell_sum

            doc_sheet.cell(row=doc_row, column=6).value = f'{cell_sum:.2f}'
            doc_sheet.cell(row=doc_row, column=6).font = cell_font
            doc_sheet.cell(row=doc_row, column=6).fill = cell_color 
            doc_sheet.cell(row=doc_row, column=6).alignment = end_alignment 

        doc_sheet.cell(row=2, column=2).value = f'Итого: {total_sum:.2f}'
        doc_sheet.cell(row=2, column=2).font = Font(size=18.0, color='ffffff')
